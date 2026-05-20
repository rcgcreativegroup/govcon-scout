import argparse
import csv
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from pypdf import PdfReader

try:
    import openpyxl
except ImportError:
    openpyxl = None


DEFAULT_DOWNLOADS_DIR = "downloads"
DEFAULT_EXTRACTS_DIR = "reports/document_extracts"
DEFAULT_REVIEWS_DIR = "reports/opportunity_reviews"
DEFAULT_CSV_PATH = "exports/govcon_scout_opportunities_latest.csv"

SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".txt",
    ".csv",
    ".xlsx",
    ".docx",
}

KEYWORD_SECTIONS = {
    "submission_instructions": [
        "submit",
        "submission",
        "quote shall",
        "proposal shall",
        "offers are due",
        "quotations are due",
        "email",
        "sam.gov",
        "piee",
    ],
    "deadlines": [
        "due",
        "deadline",
        "no later than",
        "questions",
        "rfi",
        "site visit",
    ],
    "scope": [
        "performance work statement",
        "pws",
        "scope",
        "contractor shall",
        "integrated pest management",
        "ipm",
        "services include",
    ],
    "pricing": [
        "pricing",
        "price",
        "clin",
        "schedule",
        "firm fixed price",
        "ffp",
        "unit price",
        "extended price",
    ],
    "compliance": [
        "wage determination",
        "service contract",
        "insurance",
        "certification",
        "representations",
        "clauses",
        "far",
        "dfars",
    ],
    "staffing": [
        "staff",
        "personnel",
        "hours",
        "on-site",
        "onsite",
        "contractor employee",
        "qualified",
        "supervisor",
    ],
    "evaluation": [
        "evaluation",
        "award",
        "lowest price",
        "technically acceptable",
        "lpta",
        "best value",
        "past performance",
    ],
}


def safe_text(value):
    if value is None:
        return ""
    return str(value).strip()


def make_safe_name(value):
    text = safe_text(value) or "unknown"
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text)
    return text.strip("_")[:120] or "unknown"


def ensure_dir(path):
    Path(path).mkdir(parents=True, exist_ok=True)


def compact_text(text):
    text = safe_text(text)
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def read_pdf(path):
    reader = PdfReader(str(path))
    pages = []

    for index, page in enumerate(reader.pages, start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception as error:
            page_text = f"[Could not extract page {index}: {error}]"

        if page_text.strip():
            pages.append(f"\n\n--- Page {index} ---\n\n{page_text}")

    return compact_text("\n".join(pages))


def read_txt(path):
    try:
        return compact_text(path.read_text(encoding="utf-8", errors="ignore"))
    except Exception:
        return compact_text(path.read_text(errors="ignore"))


def read_csv_text(path):
    lines = []

    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as file:
        reader = csv.reader(file)

        for row in reader:
            lines.append(" | ".join(safe_text(cell) for cell in row))

    return compact_text("\n".join(lines))


def read_xlsx(path):
    if openpyxl is None:
        return "[openpyxl is not installed. Run: pip install openpyxl]"

    workbook = openpyxl.load_workbook(path, data_only=True)
    lines = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        lines.append(f"\n\n--- Sheet: {sheet_name} ---\n")

        for row in sheet.iter_rows(values_only=True):
            values = [safe_text(cell) for cell in row if safe_text(cell)]
            if values:
                lines.append(" | ".join(values))

    return compact_text("\n".join(lines))


def read_docx(path):
    """
    Lightweight DOCX reader using stdlib zip/xml.
    """
    lines = []

    with zipfile.ZipFile(path) as docx:
        xml_bytes = docx.read("word/document.xml")

    root = ET.fromstring(xml_bytes)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

    for paragraph in root.findall(".//w:p", namespace):
        texts = []

        for node in paragraph.findall(".//w:t", namespace):
            if node.text:
                texts.append(node.text)

        line = "".join(texts).strip()
        if line:
            lines.append(line)

    return compact_text("\n".join(lines))


def extract_text_from_file(path):
    extension = path.suffix.lower()

    if extension == ".pdf":
        return read_pdf(path)

    if extension == ".txt":
        return read_txt(path)

    if extension == ".csv":
        return read_csv_text(path)

    if extension == ".xlsx":
        return read_xlsx(path)

    if extension == ".docx":
        return read_docx(path)

    return ""


def classify_file(path):
    name = path.name.lower()

    if "pws" in name or "sow" in name or "performance_work" in name:
        return "PWS/SOW"

    if "pricing" in name or "price" in name or "clin" in name:
        return "Pricing"

    if "sol" in name or "solicitation" in name or "rfq" in name or "rfp" in name:
        return "Solicitation"

    if "map" in name or "footprint" in name:
        return "Map / Facility Reference"

    if "facility" in name or "facilities" in name:
        return "Facility Information"

    if "amend" in name or "sf30" in name:
        return "Amendment"

    if "sf1449" in name:
        return "SF1449"

    return "Other"


def get_supported_files(folder):
    folder = Path(folder)

    if not folder.exists():
        return []

    files = []

    for path in sorted(folder.iterdir()):
        if not path.is_file():
            continue

        if path.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(path)

    return files


def find_keyword_hits(text, max_hits_per_section=12):
    lower_text = text.lower()
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    results = {}

    for section, keywords in KEYWORD_SECTIONS.items():
        hits = []

        for line in lines:
            lower_line = line.lower()

            if any(keyword in lower_line for keyword in keywords):
                hits.append(line[:500])

            if len(hits) >= max_hits_per_section:
                break

        results[section] = hits

    return results


def load_opportunity_from_csv(notice_id, csv_path):
    path = Path(csv_path)

    if not path.exists():
        return {}

    with open(path, "r", encoding="utf-8", newline="") as file:
        reader = csv.DictReader(file)

        for row in reader:
            if row.get("notice_id") == notice_id:
                return dict(row)

    return {}


def write_extract_file(output_dir, source_file, text):
    ensure_dir(output_dir)

    output_path = Path(output_dir) / f"{source_file.stem}.txt"
    output_path.write_text(text, encoding="utf-8")

    return output_path


def summarize_document_inventory(files):
    lines = []

    for file_path in files:
        lines.append(
            f"- **{file_path.name}** — {classify_file(file_path)} — {file_path.stat().st_size:,} bytes"
        )

    return lines


def build_review_markdown(notice_id, opportunity, files, extracted_records, review_path):
    title = opportunity.get("title") or notice_id
    agency = opportunity.get("department_ind_agency", "")
    deadline = opportunity.get("due_date_user_local") or opportunity.get("response_deadline", "")
    fit_score = opportunity.get("fit_score", "")
    prime_reality = opportunity.get("prime_reality_score", "")
    recommendation = opportunity.get("conditional_recommendation") or opportunity.get("recommendation", "")
    sam_link = opportunity.get("ui_link", "")

    combined_text = "\n\n".join(record["text"] for record in extracted_records)
    hits = find_keyword_hits(combined_text)

    has_pws = any(record["type"] == "PWS/SOW" for record in extracted_records)
    has_pricing = any(record["type"] == "Pricing" for record in extracted_records)
    has_solicitation = any(record["type"] == "Solicitation" for record in extracted_records)

    readiness = "Ready" if has_pws and has_pricing else "Partial"

    lines = []
    lines.append(f"# Bid/No-Bid Review Packet — {notice_id}")
    lines.append("")
    lines.append(f"**Title:** {title}")
    lines.append(f"**Agency:** {agency}")
    lines.append(f"**Deadline:** {deadline}")
    lines.append(f"**Fit Score:** {fit_score}")
    lines.append(f"**Prime Reality Score:** {prime_reality}")
    lines.append(f"**Recommendation:** {recommendation}")
    lines.append(f"**SAM.gov Link:** {sam_link}")
    lines.append("")
    lines.append("## Readiness")
    lines.append("")
    lines.append(f"- **Document Readiness:** {readiness}")
    lines.append(f"- **PWS/SOW Found:** {'Yes' if has_pws else 'No'}")
    lines.append(f"- **Pricing File Found:** {'Yes' if has_pricing else 'No'}")
    lines.append(f"- **Solicitation File Found:** {'Yes' if has_solicitation else 'No'}")
    lines.append("")
    lines.append("## Local Document Inventory")
    lines.append("")
    lines.extend(summarize_document_inventory(files))
    lines.append("")
    lines.append("## Early Bid/No-Bid Takeaway")
    lines.append("")

    if has_pws and has_pricing:
        lines.append(
            "This opportunity is ready for deeper bid/no-bid analysis because the package includes a PWS/SOW and pricing file."
        )
    else:
        lines.append(
            "This package is partially ready. Review the extracted files and confirm whether the missing solicitation, PWS, or pricing requirements are contained inside another document."
        )

    lines.append("")
    lines.append("## Scope / PWS Clues")
    lines.append("")
    add_hits(lines, hits.get("scope", []))

    lines.append("")
    lines.append("## Submission Instructions Clues")
    lines.append("")
    add_hits(lines, hits.get("submission_instructions", []))

    lines.append("")
    lines.append("## Deadline / RFI / Site Visit Clues")
    lines.append("")
    add_hits(lines, hits.get("deadlines", []))

    lines.append("")
    lines.append("## Pricing Clues")
    lines.append("")
    add_hits(lines, hits.get("pricing", []))

    lines.append("")
    lines.append("## Compliance / Clauses Clues")
    lines.append("")
    add_hits(lines, hits.get("compliance", []))

    lines.append("")
    lines.append("## Staffing / Execution Clues")
    lines.append("")
    add_hits(lines, hits.get("staffing", []))

    lines.append("")
    lines.append("## Evaluation / Award Clues")
    lines.append("")
    add_hits(lines, hits.get("evaluation", []))

    lines.append("")
    lines.append("## Suggested Next Analysis Prompt")
    lines.append("")
    lines.append(
        "Analyze the extracted solicitation package for bid/no-bid, compliance checklist, required forms, pricing strategy, RFI questions, and a proposal outline. Focus on whether JPTR/RCG should prime, subcontract, or pass."
    )
    lines.append("")
    lines.append("## Extracted Text Files")
    lines.append("")

    for record in extracted_records:
        lines.append(f"- {record['extract_path']}")

    review_path.write_text("\n".join(lines), encoding="utf-8")


def add_hits(lines, hits):
    if not hits:
        lines.append("- No obvious clues detected in extracted text.")
        return

    for hit in hits:
        lines.append(f"- {hit}")


def extract_notice_documents(notice_id, downloads_dir, extracts_dir, reviews_dir, csv_path):
    notice_id = make_safe_name(notice_id)
    source_folder = Path(downloads_dir) / notice_id
    extract_folder = Path(extracts_dir) / notice_id
    review_folder = Path(reviews_dir)

    ensure_dir(extract_folder)
    ensure_dir(review_folder)

    files = get_supported_files(source_folder)

    if not files:
        print(f"No supported files found in: {source_folder}")
        return ""

    opportunity = load_opportunity_from_csv(notice_id, csv_path)
    extracted_records = []

    print(f"Extracting {len(files)} file(s) from {source_folder}")

    for file_path in files:
        print(f"- Extracting: {file_path.name}")

        try:
            text = extract_text_from_file(file_path)
        except Exception as error:
            text = f"[Extraction failed for {file_path.name}: {error}]"

        text = compact_text(text)

        if not text:
            text = f"[No extractable text found for {file_path.name}]"

        extract_path = write_extract_file(
            output_dir=extract_folder,
            source_file=file_path,
            text=text,
        )

        extracted_records.append({
            "file": str(file_path),
            "name": file_path.name,
            "type": classify_file(file_path),
            "extract_path": str(extract_path),
            "text": text,
        })

    review_path = review_folder / f"{notice_id}_bid_no_bid.md"

    build_review_markdown(
        notice_id=notice_id,
        opportunity=opportunity,
        files=files,
        extracted_records=extracted_records,
        review_path=review_path,
    )

    print("")
    print(f"Document extracts written to: {extract_folder}")
    print(f"Bid/no-bid review packet: {review_path}")
    print("")

    return str(review_path)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract text from local GovCon Scout downloaded documents."
    )

    parser.add_argument(
        "--notice-id",
        required=True,
        help="Notice ID folder to extract, e.g. HE125426QE041.",
    )

    parser.add_argument(
        "--downloads-dir",
        default=DEFAULT_DOWNLOADS_DIR,
        help="Folder containing downloads/{notice_id}/.",
    )

    parser.add_argument(
        "--extracts-dir",
        default=DEFAULT_EXTRACTS_DIR,
        help="Output folder for extracted text files.",
    )

    parser.add_argument(
        "--reviews-dir",
        default=DEFAULT_REVIEWS_DIR,
        help="Output folder for bid/no-bid markdown packets.",
    )

    parser.add_argument(
        "--csv",
        default=DEFAULT_CSV_PATH,
        help="GovCon Scout CSV for opportunity metadata.",
    )

    return parser.parse_args()


def main():
    args = parse_args()

    extract_notice_documents(
        notice_id=args.notice_id,
        downloads_dir=args.downloads_dir,
        extracts_dir=args.extracts_dir,
        reviews_dir=args.reviews_dir,
        csv_path=args.csv,
    )


if __name__ == "__main__":
    main()